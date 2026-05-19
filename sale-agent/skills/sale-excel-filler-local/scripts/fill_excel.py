import openpyxl
import sys
import io
from openpyxl.styles import Border, Side, PatternFill, Alignment

# Fix Windows console encoding
if sys.platform == 'win32':
    try:
        sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
    except:
        pass

def fill_excel(file_path: str, data: dict) -> dict:
    """
    填写Excel文件（支持自动轮次：初试/线下协同初试/复试/终试）
    Args:
        file_path: Excel文件路径
        data: 包含resume_data和interview_data的字典
    Returns:
        写入的字段列表
    """
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active

    # ========== 关键修复1：先保存所有需要恢复的单元格填充色（合并前） ==========
    # 保存合并单元格区域的填充色（覆盖所有关键高亮区域）
    fill_color_map = {}
    # 亮点/疑点/风险点区域
    for row in [10, 11, 12]:
        for col in ['A', 'B', 'C', 'D', 'E', 'F']:
            cell = f"{col}{row}"
            fill = ws[cell].fill
            if fill and fill.fill_type == 'solid':
                fill_color_map[cell] = fill.fgColor.rgb
    # G10-I10区域
    for col in ['G', 'H', 'I']:
        cell = f"{col}10"
        fill = ws[cell].fill
        if fill and fill.fill_type == 'solid':
            fill_color_map[cell] = fill.fgColor.rgb
    # J10-L10区域（原脚本遗漏）
    for col in ['J', 'K', 'L']:
        cell = f"{col}10"
        fill = ws[cell].fill
        if fill and fill.fill_type == 'solid':
            fill_color_map[cell] = fill.fgColor.rgb

    # 保存原始合并范围（仅用于后续恢复非关键区域）
    original_merged_ranges = list(ws.merged_cells.ranges)

    # ========== 取消合并（仅必要区域） ==========
    to_unmerge = ['A3:C3', 'G3:I3', 'A4:F4', 'G4:I4', 'A10:F10', 'A11:F11', 'A12:F12', 'G10:I10', 'J10:L10', 'A40:B43', 'G42:I42', 'J42:L42']
    for cell_range in to_unmerge:
        try:
            ws.unmerge_cells(cell_range)
        except Exception:
            pass

    filled_fields = []

    # ========== 填写简历信息 ==========
    resume_data = data.get("resume_data", {})
    if resume_data:
        # 基础信息
        if resume_data.get("姓名"):
            ws["B3"] = resume_data["姓名"]
            filled_fields.append("B3: 姓名")
        if resume_data.get("籍贯"):
            ws["D3"] = resume_data["籍贯"]
            filled_fields.append("D3: 籍贯")
        if resume_data.get("出生年月"):
            ws["F3"] = resume_data["出生年月"]
            filled_fields.append("F3: 出生年月")
        if resume_data.get("应聘岗位"):
            ws["B4"] = resume_data["应聘岗位"]
            filled_fields.append("B4: 应聘岗位")
        if resume_data.get("意向城市"):
            ws["D4"] = resume_data["意向城市"]
            filled_fields.append("D4: 意向城市")
        if resume_data.get("原公司类型"):
            ws["F4"] = resume_data["原公司类型"]
            filled_fields.append("F4: 原公司类型")
        if resume_data.get("打败的竞争对手"):
            ws["G4"] = resume_data["打败的竞争对手"]
            filled_fields.append("G4: 打败的竞争对手")
        if resume_data.get("第一学历学校"):
            ws["B5"] = resume_data["第一学历学校"]
            filled_fields.append("B5: 第一学历学校")
        if resume_data.get("专业"):
            ws["D5"] = resume_data["专业"]
            filled_fields.append("D5: 专业")
        if resume_data.get("底薪/绩效/奖金"):
            ws["E5"] = resume_data["底薪/绩效/奖金"]
            filled_fields.append("E5: 底薪/绩效/奖金")

        # 人才类型（F6）
        if resume_data.get("人才类型"):
            ws["F6"] = resume_data["人才类型"]
            filled_fields.append("F6: 人才类型")

        # 离职原因（B7）
        if resume_data.get("离职原因"):
            ws["B7"] = resume_data["离职原因"]
            filled_fields.append("B7: 离职原因")

        # HR/人才来源（F7，明确标注：E7是标签，F7填值）
        if resume_data.get("HR/人才来源"):
            ws["F7"] = resume_data["HR/人才来源"]
            filled_fields.append("F7: HR/人才来源")

        # 亮点/疑点/风险点（仅简历填写，前缀强制添加）
        if resume_data.get("亮点"):
            ws["A10"] = "亮点：" + resume_data["亮点"]
            filled_fields.append("A10: 亮点")
        if resume_data.get("疑点"):
            ws["A11"] = "疑点：" + resume_data["疑点"]
            filled_fields.append("A11: 疑点")
        if resume_data.get("风险点"):
            ws["A12"] = "风险点：" + resume_data["风险点"]
            filled_fields.append("A12: 风险点")

    # ========== 填写面试信息（支持：初试/线下协同初试/复试/终试） ==========
    interview_data = data.get("interview_data", {})
    if interview_data:
        round_type = interview_data.get("round_type", "初试")

        # 轮次列映射（严格按SKILL.md规则）
        round_mapping = {
            "初试": {"score_col": "D", "basis_col": "E", "result_col": "D", "rating_col": "E"},
            "线下协同初试": {"score_col": "G", "basis_col": "H", "result_col": "G", "rating_col": "H"},
            "复试": {"score_col": "J", "basis_col": "K", "result_col": "J", "rating_col": "K"},
            "终试": {"score_col": "M", "basis_col": "N", "result_col": "M", "rating_col": "N"},
        }
        # 默认初试
        col_mapping = round_mapping.get(round_type, round_mapping["初试"])
        score_col = col_mapping["score_col"]
        basis_col = col_mapping["basis_col"]
        result_col = col_mapping["result_col"]
        rating_col = col_mapping["rating_col"]

        # 评分项映射（行号严格对应SKILL.md，包含单元三板斧强制填写）
        score_mapping = {
            # 底色
            "底色-兴趣": 17,
            "底色-主动侵略性": 18,
            "底色-目标刚性": 19,
            "底色-结果闭环": 20,
            "底色-信任构建能力": 21,
            "底色-人性洞察能力": 22,
            "底色-长期价值绑定": 23,
            # 岗位三板斧
            "岗位三板斧-搞成铁杆能力": 24,
            "岗位三板斧-铁杆数量": 25,
            "岗位三板斧-铁杆质量": 26,
            "岗位三板斧-信息来源": 27,
            "岗位三板斧-信息收集和整理能力": 28,
            "岗位三板斧-情报质量": 29,
            "岗位三板斧-产品知识": 30,
            "岗位三板斧-行业知识": 31,
            "岗位三板斧-输出质量": 32,
            # 单元三板斧（所有岗位强制填写）
            "单元三板斧-画施工图": 33,
            "单元三板斧-选好骨干": 34,
            "单元三板斧-提高人效": 35,
        }

        basis_data = interview_data.get("评分依据", {})

        # 字段名兼容映射（sale-IID输出 → Excel字段名）
        field_aliases = {
            "底色-狼性-主动侵略性": "底色-主动侵略性",
            "底色-狼性-目标刚性": "底色-目标刚性",
            "底色-狼性-结果闭环": "底色-结果闭环",
            "底色-情商-信任构建能力": "底色-信任构建能力",
            "底色-情商-人性洞察能力": "底色-人性洞察能力",
            "底色-情商-长期价值绑定": "底色-长期价值绑定",
        }

        # 填写评分和依据（跳过AI能力字段）
        for field, row in score_mapping.items():
            # 跳过AI能力（SKILL.md明确不填写）
            if "AI能力" in field:
                continue
            
            # 先匹配原字段，再匹配别名
            value = interview_data.get(field)
            basis = basis_data.get(field)
            if value is None:
                alias = field_aliases.get(field)
                if alias:
                    value = interview_data.get(alias)
                    basis = basis_data.get(alias)
            
            # 填写评分
            if value:
                ws[f"{score_col}{row}"] = value
                filled_fields.append(f"{score_col}{row}: {field}")
            # 填写依据
            if basis:
                ws[f"{basis_col}{row}"] = basis
                filled_fields.append(f"{basis_col}{row}: {field}依据")

        # 面试结果（行37，SKILL.md明确：行36/38不填）
        if interview_data.get("面试结果"):
            ws[f"{result_col}37"] = interview_data["面试结果"]
            filled_fields.append(f"{result_col}37: 面试结果")

        # 面试评级（行37，对应评级列）
        if interview_data.get("人才类型"):
            ws[f"{rating_col}37"] = interview_data["人才类型"]
            filled_fields.append(f"{rating_col}37: 面试评级")

    # ========== 恢复合并与样式 ==========
    thin = Side(style='thin', color='FF000000')

    # 步骤1：重新合并关键单元格（先合并，再恢复样式）
    merge_cells_list = [
        'A10:F10', 'A11:F11', 'A12:F12', 
        'G10:I10', 'J10:L10', 
        'A40:B43', 'G42:I42', 'J42:L42'
    ]
    for cell_range in merge_cells_list:
        try:
            ws.merge_cells(cell_range)
        except Exception as e:
            pass  # 已合并则忽略

    # 步骤2：恢复原始合并范围（非关键区域）
    skip_ranges = ['A10:F10', 'A11:F11', 'A12:F12', 'G10:I10', 'J10:L10']
    for merged_range in original_merged_ranges:
        range_str = str(merged_range)
        if any(skip in range_str for skip in skip_ranges):
            continue
        try:
            ws.merge_cells(range_str)
        except:
            pass

    # ========== 关键修复2：恢复所有保存的填充色 ==========
    for cell, rgb in fill_color_map.items():
        try:
            ws[cell].fill = PatternFill(fill_type='solid', fgColor=rgb)
        except Exception as e:
            pass

    # 自动换行（保证长文本显示）
    wrap_alignment = Alignment(wrap_text=True, vertical='top', horizontal='left')
    ws['A10'].alignment = wrap_alignment
    ws['A11'].alignment = wrap_alignment
    ws['A12'].alignment = wrap_alignment

    # 边框样式（仅补充缺失边框，不覆盖原有样式）
    ws['F10'].border = Border(right=thin, top=thin, bottom=thin)
    ws['F11'].border = Border(right=thin, top=thin, bottom=thin)
    ws['F12'].border = Border(right=thin, top=thin, bottom=thin)
    ws['G10'].border = Border(left=thin, top=thin, bottom=thin)

    # 行高设置（固定行高）
    ws.row_dimensions[10].height = 36
    ws.row_dimensions[11].height = 36
    ws.row_dimensions[12].height = 36
    
    # 保存并关闭文件
    wb.save(file_path)
    wb.close()

    return {"filled_fields": filled_fields}


def main():
    """命令行入口"""
    import json
    import sys

    if len(sys.argv) < 3:
        print("使用方法: python fill_excel.py <Excel文件.xlsx> <数据文件.json> [轮次]")
        print("示例: python fill_excel.py template.xlsx data.json 线下协同初试")
        sys.exit(1)

    file_path = sys.argv[1]
    json_path = sys.argv[2]
    round_type = sys.argv[3] if len(sys.argv) > 3 else "初试"

    # 读取JSON数据
    with open(json_path, "r", encoding="utf-8") as f:
        data = json.load(f)

    # 设置轮次（优先级：命令行 > JSON内配置）
    if "interview_data" not in data:
        data["interview_data"] = {}
    data["interview_data"]["round_type"] = round_type

    # 填写Excel
    result = fill_excel(file_path, data)
    print(f"✅ 填写完成，已填写字段数: {len(result['filled_fields'])}")
    for field in result["filled_fields"]:
        print(f"  - {field}")


if __name__ == "__main__":
    main()