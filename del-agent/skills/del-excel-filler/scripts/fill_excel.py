import openpyxl
import sys
import io
from openpyxl.styles import Border, Side, PatternFill, Alignment
from openpyxl.utils import get_column_letter
import json

# Fix Windows console encoding
if sys.platform == 'win32':
    try:
        sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
    except Exception as e:
        print(f"控制台编码修复失败: {e}", file=sys.stderr)

def fill_excel(file_path: str, data: dict) -> dict:
    """
    填写交付岗位Excel文件（支持自动轮次：初试/线下协同初试/复试/终试）
    修复点：
    1. 移除电话字段填写逻辑
    2. 修复单元格消失问题（合并/取消合并+样式丢失）
    3. 增强空值处理，避免数据缺失
    4. 修复合并单元格赋值逻辑
    5. 统一单元格样式，避免显示异常
    6. 增加字段映射容错性
    Args:
        file_path: Excel文件路径
        data: 包含resume_data和interview_data的字典
    Returns:
        写入的字段列表
    """
    # 容错：文件不存在/损坏时抛出明确异常
    try:
        wb = openpyxl.load_workbook(file_path)
    except FileNotFoundError:
        raise FileNotFoundError(f"Excel文件不存在: {file_path}")
    except Exception as e:
        raise Exception(f"加载Excel文件失败: {e}")

    ws = wb.active
    filled_fields = []

    # ========== 1. 初始化样式配置 ==========
    thin_border = Side(style='thin', color='FF000000')
    default_alignment = Alignment(
        wrap_text=True,
        vertical='center',
        horizontal='left'
    )
    # 默认填充色（避免透明/无填充导致单元格"消失"）
    default_fill = PatternFill(fill_type='solid', fgColor='FFFFFF')

    # ========== 2. 保存原始样式（全量保存，避免丢失） ==========
    fill_color_map = {}
    border_map = {}
    alignment_map = {}

    # 保存关键区域样式：基础信息+面试评分+建议（移除C3/电话字段）
    key_cells = [
        # 基础信息（移除C3）
        'B3', 'D3', 'F3', 'B4', 'D4', 'F4', 'B5', 'D5', 'F5',
        'A10', 'A11', 'A12',
        # 面试评分
        *[f"{col}{row}" for col in ['D','E','G','H','J','K','M','N'] for row in range(16, 31)],
        # 建议字段
        *[f"{col}{row}" for col in ['C','G','J','M'] for row in range(33, 37)]
    ]

    for cell in key_cells:
        try:
            # 保存填充色
            fill = ws[cell].fill
            if fill and fill.fill_type == 'solid' and fill.fgColor.rgb:
                fill_color_map[cell] = fill.fgColor.rgb
            else:
                fill_color_map[cell] = 'FFFFFF'  # 白色兜底

            # 保存边框
            border = ws[cell].border
            if border and all([border.left, border.right, border.top, border.bottom]):
                border_map[cell] = border
            else:
                border_map[cell] = Border(
                    left=thin_border, right=thin_border,
                    top=thin_border, bottom=thin_border
                )
            
            # 保存对齐方式
            alignment_map[cell] = ws[cell].alignment or default_alignment
        except Exception as e:
            print(f"保存单元格{cell}样式失败: {e}", file=sys.stderr)
            # 兜底样式
            fill_color_map[cell] = 'FFFFFF'
            border_map[cell] = Border(left=thin_border, right=thin_border, top=thin_border, bottom=thin_border)
            alignment_map[cell] = default_alignment

    # ========== 3. 安全取消合并单元格（避免残留MergedCell） ==========
    # 先获取所有已合并的单元格范围，再针对性取消
    merged_ranges = list(ws.merged_cells.ranges)
    to_unmerge = [
        'B3:C3', 'D3:F3', 'B4:F4',
        'A10:F10', 'A11:F11', 'A12:F12',
        'C33:C36', 'G33:G36', 'J33:J36', 'M33:M36'
    ]
    for cell_range in to_unmerge:
        try:
            # 先检查是否真的合并了，再取消
            if any(str(rng) == cell_range for rng in merged_ranges):
                ws.unmerge_cells(cell_range)
        except Exception as e:
            print(f"取消合并{cell_range}失败: {e}", file=sys.stderr)
            pass
    
    # 额外：取消评分区域每行的合并（模板中部分行有跨列合并）
    for col in ['D', 'E', 'G', 'H', 'J', 'K', 'M', 'N']:
        for row in range(16, 31):
            cell_range = f'{col}{row}'
            try:
                if any(str(rng) == cell_range for rng in merged_ranges):
                    ws.unmerge_cells(cell_range)
            except Exception:
                pass

    # ========== 4. 填写简历信息（移除电话字段） ==========
    resume_data = data.get("resume_data", {})
    if resume_data:
        # 基础信息 - 逐个字段兜底，避免空值跳过（移除电话/C3）
        field_mapping = {
            "姓名": "B3",
            "籍贯": "D3",
            "应聘岗位": "B4",
            "意向城市": "D4",
            "原公司类型": "F4",
            "第一学历学校": "B5",
            "专业": "D5",
            "高考数学成绩": "F5"
        }

        for field_name, cell in field_mapping.items():
            value = resume_data.get(field_name, "").strip()  # 去除首尾空格
            if value or value == "0":  # 兼容0值（如数学成绩0分）
                try:
                    ws[cell] = value
                    filled_fields.append(f"{cell}: {field_name}")
                except Exception as e:
                    print(f"填写{cell}({field_name})失败: {e}", file=sys.stderr)
                    pass
            
            # 强制恢复样式（核心：避免单元格消失）
            try:
                ws[cell].alignment = alignment_map.get(cell, default_alignment)
                ws[cell].fill = PatternFill(fill_type='solid', fgColor=fill_color_map.get(cell, 'FFFFFF'))
                ws[cell].border = border_map.get(cell)
            except Exception as e:
                print(f"恢复{cell}样式失败: {e}", file=sys.stderr)
                pass

        # 出生年月+民族（合并单元格F3）- 兼容单字段为空的情况
        birth = resume_data.get("出生年月", "").strip()
        nation = resume_data.get("民族", "").strip()
        birth_nation = []
        if birth:
            birth_nation.append(birth)
        if nation:
            birth_nation.append(nation)
        # 即使都为空，也初始化单元格
        try:
            ws["F3"] = " / ".join(birth_nation) if birth_nation else ""
            filled_fields.append("F3: 出生年月/民族")
        except Exception as e:
            print(f"填写F3(出生年月/民族)失败: {e}", file=sys.stderr)
            pass
        # 强制恢复F3样式
        try:
            ws["F3"].alignment = alignment_map.get("F3", default_alignment)
            ws["F3"].fill = PatternFill(fill_type='solid', fgColor=fill_color_map.get("F3", 'FFFFFF'))
            ws["F3"].border = border_map.get("F3")
        except Exception as e:
            print(f"恢复F3样式失败: {e}", file=sys.stderr)
            pass

        # 亮点/疑点/风险点 - 兼容空值，统一前缀和样式
        remark_fields = {
            "亮点": "A10",
            "疑点": "A11",
            "风险点": "A12"
        }
        for field_name, cell in remark_fields.items():
            value = resume_data.get(field_name, "").strip()
            display_value = f"{field_name}：{value}" if value else f"{field_name}：无"
            try:
                ws[cell] = display_value
                filled_fields.append(f"{cell}: {field_name}")
            except Exception as e:
                print(f"填写{cell}({field_name})失败: {e}", file=sys.stderr)
                pass
            # 强制设置样式（避免单元格消失）
            try:
                # 强制设置自动换行和行高
                ws[cell].alignment = Alignment(wrap_text=True, vertical='top', horizontal='left')
                ws.row_dimensions[int(cell[1:])].height = 36
                # 恢复背景色
                ws[cell].fill = PatternFill(fill_type='solid', fgColor=fill_color_map.get(cell, 'FFFFE0'))
                # 补全边框
                ws[cell].border = Border(
                    left=thin_border, right=thin_border,
                    top=thin_border, bottom=thin_border
                )
            except Exception as e:
                print(f"恢复{cell}样式失败: {e}", file=sys.stderr)
                pass

    # ========== 5. 填写面试信息（增强容错性） ==========
    interview_data = data.get("interview_data", {})
    if interview_data:
        # 轮次映射容错 - 兼容大小写/错别字
        round_type = interview_data.get("round_type", "初试").strip()
        round_type_normalized = round_type.lower()
        round_mapping = {
            "初试": {"score_col": "D", "basis_col": "E", "suggest_col": "C"},
            "线下协同初试": {"score_col": "G", "basis_col": "H", "suggest_col": "G"},
            "复试": {"score_col": "J", "basis_col": "K", "suggest_col": "J"},
            "终试": {"score_col": "M", "basis_col": "N", "suggest_col": "M"},
        }

        # 轮次自动修正
        if "初试" in round_type_normalized or "一面" in round_type_normalized:
            col_mapping = round_mapping["初试"]
        elif "线下协同" in round_type_normalized:
            col_mapping = round_mapping["线下协同初试"]
        elif "复试" in round_type_normalized or "二面" in round_type_normalized:
            col_mapping = round_mapping["复试"]
        elif "终试" in round_type_normalized or "三面" in round_type_normalized:
            col_mapping = round_mapping["终试"]
        else:
            col_mapping = round_mapping["初试"]  # 默认

        score_col = col_mapping["score_col"]
        basis_col = col_mapping["basis_col"]
        suggest_col = col_mapping["suggest_col"]

        # 评分项映射 - 全量覆盖，避免遗漏
        score_mapping = {
            "AI能力": 16,
            "底色-兴趣": 17,
            "底色-智商-数学成绩": 18,
            "底色-智商-毕业院校": 19,
            "底色-智商-简化能力": 20,
            "底色-狼性": 21,
            "底色-情商": 22,
            "岗位三板斧-硬能力": 23,
            "岗位三板斧-软能力": 24,
            "岗位三板斧-态度": 25,
            "单元三板斧-画施工图": 26,
            "单元三板斧-选好骨干": 27,
            "单元三板斧-提高人效": 28,
            "结果-总分": 30,
            "面试评级": 30
        }

        basis_data = interview_data.get("评分依据", {})

        # 填写评分和依据 - 兼容空值和非数值类型
        for field, row in score_mapping.items():
            try:
                if field == "面试评级":
                    # 面试评级填写到依据列第30行
                    value = interview_data.get(field, "").strip()
                    if value:
                        try:
                            ws[f"{basis_col}{row}"] = value
                            filled_fields.append(f"{basis_col}{row}: {field}")
                        except Exception as e:
                            print(f"填写{basis_col}{row}({field})失败: {e}", file=sys.stderr)
                            pass
                    # 初始化样式
                    try:
                        cell = f"{basis_col}{row}"
                        ws[cell].alignment = alignment_map.get(cell, default_alignment)
                        ws[cell].fill = PatternFill(fill_type='solid', fgColor=fill_color_map.get(cell, 'FFFFFF'))
                        ws[cell].border = border_map.get(cell)
                    except Exception as e:
                        print(f"恢复{cell}样式失败: {e}", file=sys.stderr)
                        pass
                    continue

                # 填写评分值 - 兼容数值/字符串
                value = interview_data.get(field)
                if value is not None and value != "":
                    try:
                        # 转换为字符串避免类型错误
                        ws[f"{score_col}{row}"] = str(value).strip()
                        filled_fields.append(f"{score_col}{row}: {field}")
                    except Exception as e:
                        print(f"填写{score_col}{row}({field})失败: {e}", file=sys.stderr)
                        pass

                # 填写评分依据
                basis = basis_data.get(field, "").strip()
                if basis:
                    try:
                        ws[f"{basis_col}{row}"] = basis
                        filled_fields.append(f"{basis_col}{row}: {field}依据")
                    except Exception as e:
                        print(f"填写{basis_col}{row}({field}依据)失败: {e}", file=sys.stderr)
                        pass

                # 统一初始化样式（核心：避免单元格消失）
                score_cell = f"{score_col}{row}"
                basis_cell = f"{basis_col}{row}"
                # 评分单元格样式
                try:
                    ws[score_cell].alignment = alignment_map.get(score_cell, default_alignment)
                    ws[score_cell].fill = PatternFill(fill_type='solid', fgColor=fill_color_map.get(score_cell, 'FFFFFF'))
                    ws[score_cell].border = border_map.get(score_cell)
                except Exception as e:
                    print(f"恢复{score_cell}样式失败: {e}", file=sys.stderr)
                    pass
                # 依据单元格样式
                try:
                    ws[basis_cell].alignment = alignment_map.get(basis_cell, default_alignment)
                    ws[basis_cell].fill = PatternFill(fill_type='solid', fgColor=fill_color_map.get(basis_cell, 'FFFFFF'))
                    ws[basis_cell].border = border_map.get(basis_cell)
                except Exception as e:
                    print(f"恢复{basis_cell}样式失败: {e}", file=sys.stderr)
                    pass
            except Exception as e:
                print(f"填写评分项{field}失败: {e}", file=sys.stderr)
                continue

        # 面试建议字段 - 兼容空值，统一样式
        suggest_mapping = {
            "级别建议": 33,
            "薪资建议": 34,
            "说服建议": 35,
            "其他建议": 36
        }
        for field, row in suggest_mapping.items():
            try:
                value = interview_data.get(field, "").strip()
                if value or value == "0":
                    try:
                        ws[f"{suggest_col}{row}"] = value
                        filled_fields.append(f"{suggest_col}{row}: {field}")
                    except Exception as e:
                        print(f"填写{suggest_col}{row}({field})失败: {e}", file=sys.stderr)
                        pass

                # 初始化样式（核心：避免单元格消失）
                cell = f"{suggest_col}{row}"
                try:
                    ws[cell].alignment = alignment_map.get(cell, default_alignment)
                    ws[cell].fill = PatternFill(fill_type='solid', fgColor=fill_color_map.get(cell, 'FFFFFF'))
                    ws[cell].border = border_map.get(cell)
                except Exception as e:
                    print(f"恢复{cell}样式失败: {e}", file=sys.stderr)
                    pass
            except Exception as e:
                print(f"填写建议项{field}失败: {e}", file=sys.stderr)
                continue

    # ========== 6. 安全合并单元格（修复合并逻辑，避免单元格消失） ==========
    merge_cells_list = [
        'B3:C3', 'D3:F3', 'B4:F4',
        'A10:F10', 'A11:F11', 'A12:F12',
        'C33:C36', 'G33:G36', 'J33:J36', 'M33:M36'
    ]
    for cell_range in merge_cells_list:
        try:
            # 先检查是否已合并，避免重复合并
            if not any(str(rng) == cell_range for rng in ws.merged_cells.ranges):
                ws.merge_cells(cell_range)
            # 合并后强制设置样式
            start_cell = cell_range.split(':')[0]
            ws[start_cell].alignment = default_alignment
            ws[start_cell].fill = default_fill
            ws[start_cell].border = thin_border
        except Exception as e:
            print(f"合并单元格{cell_range}失败: {e}", file=sys.stderr)
            pass

    # ========== 7. 最终样式优化（避免单元格显示异常） ==========
    # 统一行高（确保内容可见）
    for row in [10, 11, 12]:
        ws.row_dimensions[row].height = 36
    # 统一列宽（避免内容被遮挡）
    for col in range(1, 15):  # A到N列
        ws.column_dimensions[get_column_letter(col)].width = 18
    
    # 强制恢复所有关键单元格的边框（防止边框丢失导致"消失"）
    for cell in key_cells:
        try:
            ws[cell].border = border_map.get(cell)
        except:
            pass

    # ========== 8. 保存文件（安全处理） ==========
    try:
        wb.save(file_path)
    except PermissionError:
        raise PermissionError(f"无权限保存文件，请关闭Excel后重试: {file_path}")
    finally:
        wb.close()

    return {"filled_fields": filled_fields}


def main():
    """命令行入口 - 增强参数校验和错误处理"""
    if len(sys.argv) < 3:
        print("""
使用方法: python fill_excel.py <Excel文件.xlsx> <数据文件.json> [轮次]
参数说明:
  <Excel文件.xlsx>  - 目标Excel模板文件路径
  <数据文件.json>   - 包含简历和面试数据的JSON文件
  [轮次]           - 可选，面试轮次（初试/线下协同初试/复试/终试）
示例:
  python fill_excel.py template.xlsx data.json 线下协同初试
        """, file=sys.stderr)
        sys.exit(1)

    file_path = sys.argv[1].strip()
    json_path = sys.argv[2].strip()
    round_type = sys.argv[3].strip() if len(sys.argv) > 3 else "初试"

    # 读取JSON数据 - 增强容错
    try:
        with open(json_path, "r", encoding="utf-8") as f:
            data = json.load(f)
    except FileNotFoundError:
        print(f"JSON文件不存在: {json_path}", file=sys.stderr)
        sys.exit(1)
    except json.JSONDecodeError:
        print(f"JSON文件格式错误: {json_path}", file=sys.stderr)
        sys.exit(1)
    except Exception as e:
        print(f"读取JSON文件失败: {e}", file=sys.stderr)
        sys.exit(1)

    # 设置轮次（优先级：命令行 > JSON内配置）
    if "interview_data" not in data:
        data["interview_data"] = {}
    data["interview_data"]["round_type"] = round_type

    # 填写Excel并捕获异常
    try:
        result = fill_excel(file_path, data)
        print(f"填写完成，已填写字段数: {len(result['filled_fields'])}")
        for idx, field in enumerate(result["filled_fields"], 1):
            print(f"  {idx}. {field}")
    except Exception as e:
        print(f"填写失败: {e}", file=sys.stderr)
        sys.exit(1)


if __name__ == "__main__":
    main()